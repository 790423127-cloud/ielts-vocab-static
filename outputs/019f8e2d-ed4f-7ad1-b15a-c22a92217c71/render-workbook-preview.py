from pathlib import Path
from textwrap import wrap

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


OUTPUT_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = OUTPUT_DIR / "雅思阅读538考点词真经-网站实际376词.xlsx"
FONT_PATH = Path("C:/Windows/Fonts/msyh.ttc")


def font(size: int, bold: bool = False):
    preferred = Path("C:/Windows/Fonts/msyhbd.ttc") if bold else FONT_PATH
    return ImageFont.truetype(str(preferred if preferred.exists() else FONT_PATH), size)


def color(cell, fallback: str) -> str:
    if not cell.fill.fill_type:
        return fallback
    value = cell.fill.fgColor.rgb
    if value and len(value) == 8 and not value.startswith("00"):
        return f"#{value[2:]}"
    return fallback


def display_value(formula_cell, value_cell) -> str:
    value = value_cell.value if formula_cell.data_type == "f" else formula_cell.value
    return "" if value is None else str(value)


def render_range(sheet, values_sheet, min_row, max_row, min_col, max_col, target):
    widths = []
    for column in range(min_col, max_col + 1):
        letter = get_column_letter(column)
        excel_width = sheet.column_dimensions[letter].width or 12
        widths.append(max(76, min(390, int(excel_width * 7.2))))

    row_heights = []
    for row in range(min_row, max_row + 1):
        excel_height = sheet.row_dimensions[row].height or 22
        row_heights.append(max(28, int(excel_height * 1.35)))

    image = Image.new("RGB", (sum(widths) + 2, sum(row_heights) + 2), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    title_font = font(20, True)
    header_font = font(13, True)
    body_font = font(12)
    bold_font = font(12, True)
    phonetic_font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 12)

    merged_anchors = {}
    merged_hidden = set()
    for merged in sheet.merged_cells.ranges:
        if (
            merged.min_row >= min_row
            and merged.max_row <= max_row
            and merged.min_col >= min_col
            and merged.max_col <= max_col
        ):
            merged_anchors[(merged.min_row, merged.min_col)] = merged
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    if (row, col) != (merged.min_row, merged.min_col):
                        merged_hidden.add((row, col))

    y = 1
    for row_offset, row in enumerate(range(min_row, max_row + 1)):
        x = 1
        for col_offset, column in enumerate(range(min_col, max_col + 1)):
            if (row, column) in merged_hidden:
                x += widths[col_offset]
                continue

            merged = merged_anchors.get((row, column))
            width = (
                sum(widths[column - min_col : merged.max_col - min_col + 1])
                if merged
                else widths[col_offset]
            )
            height = (
                sum(row_heights[row - min_row : merged.max_row - min_row + 1])
                if merged
                else row_heights[row_offset]
            )
            cell = sheet.cell(row, column)
            value_cell = values_sheet.cell(row, column)

            fallback_fill = "#FFFFFF"
            is_table_header = (
                (sheet.title == "376词表" and row == 5)
                or (sheet.title == "说明与汇总" and row == 9)
            )
            if is_table_header:
                fallback_fill = "#0F766E"
            elif row > 5 and sheet.title == "376词表":
                fallback_fill = "#F0FDFA" if row % 2 == 0 else "#FFFFFF"
            elif row > 9 and sheet.title == "说明与汇总":
                fallback_fill = "#F0FDFA" if row % 2 == 0 else "#FFFFFF"

            fill = color(cell, fallback_fill)
            draw.rectangle((x, y, x + width, y + height), fill=fill, outline="#D1D5DB", width=1)

            text = display_value(cell, value_cell)
            if row == 1:
                current_font = title_font
                text_color = "#FFFFFF"
            elif is_table_header:
                current_font = header_font
                text_color = "#FFFFFF"
            elif cell.font.bold:
                current_font = bold_font
                text_color = "#111827"
            else:
                current_font = (
                    phonetic_font
                    if sheet.title == "376词表" and column == 6 and row > 5
                    else body_font
                )
                text_color = "#1F2937"

            approximate_chars = max(5, int((width - 14) / 12))
            lines = wrap(text, width=approximate_chars, break_long_words=True, break_on_hyphens=False) or [""]
            lines = lines[:3]
            line_height = current_font.size + 5
            text_height = line_height * len(lines)
            text_y = y + max(6, (height - text_height) // 2)
            centered = row in (1, 5, 9) or column <= 4

            for line in lines:
                text_width = draw.textbbox((0, 0), line, font=current_font)[2]
                text_x = x + max(7, (width - text_width) // 2) if centered else x + 7
                draw.text((text_x, text_y), line, font=current_font, fill=text_color)
                text_y += line_height

            x += widths[col_offset]
        y += row_heights[row_offset]

    image.save(target)


workbook = load_workbook(WORKBOOK_PATH, data_only=False)
values_workbook = load_workbook(WORKBOOK_PATH, data_only=True)

assert workbook.sheetnames == ["说明与汇总", "376词表"]
assert workbook["说明与汇总"].tables["GroupSummaryTable"].ref == "A9:F17"
assert workbook["376词表"].tables["GuixueIelts376Table"].ref == "A5:I381"
assert workbook["376词表"].max_row == 381
assert workbook["376词表"]["A381"].value == 376

render_range(
    workbook["说明与汇总"],
    values_workbook["说明与汇总"],
    1,
    17,
    1,
    6,
    OUTPUT_DIR / "preview-summary.png",
)
render_range(
    workbook["376词表"],
    values_workbook["376词表"],
    1,
    25,
    1,
    9,
    OUTPUT_DIR / "preview-words.png",
)

print("Workbook structure and previews verified.")
